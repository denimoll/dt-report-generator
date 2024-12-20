# Документация по API
# https://yoursky.blue/documentation-api/dependencytrack.html

import json
from datetime import datetime

import requests
import urllib3
import validators
from docxtpl import DocxTemplate, RichText
from openpyxl import load_workbook

urllib3.disable_warnings()


def report(config):
    """create report from DT"""
    # variables
    doc = DocxTemplate("reports/draft.docx") # docx template
    excel = load_workbook("reports/draft.xlsx") # excel document
    project_info = {} # common info about project
    components = {} # dict of components 
    vulns = [] # list of vulnerabilities

    try:
        # read config
        url = config.get('url')[0]
        token = config.get('token')[0]
        project = config.get('project')[0].split("(")[1].split(")")[0]
        severities = config.get('severities')
        report_type = config.get('report_type')[0]
        
        # header for auth request
        headers = {
            'X-Api-Key': token
        }

        # format url to "protocol"://"domain"/api/v1/
        if not validators.url(url):
            raise ValueError('URL not valid')
        url = url.split('/api/v')[0] + '/api/v1/'

        # validate all parameters
        if not token:
            raise ValueError('Token not set')
        if not project:
            raise ValueError('Project not set')
        if not severities:
            raise ValueError('Severities not set')
        for i in severities:
            if i not in ['critical', 'high', 'medium', 'low', 'unassigned']:
                raise ValueError('Severities contain wrong value')
        if report_type not in ['word', 'excel']:
            raise ValueError('Report type not valid')

        # get common info about project
        res = requests.get(url+'project/'+project, headers=headers, verify=False, timeout=1000)
        if res.status_code != 200:
            raise ConnectionError('Something wrong with connection. Check your parameters')
        text = json.loads(res.text)
        project_name = RichText()
        project_name_str = text.get('name')
        project_name.add(project_name_str, url_id=doc.build_url_id(url.split('api/v1/')[0]+'projects/'+project))
        project_info.update({
            'name': project_name,
            'version': text.get('version'),
            'lastBomImport': datetime.fromtimestamp(int(text.get('lastBomImport') or 0)/1000).strftime("%d.%m.%Y %H:%M"),
            'date': datetime.now().strftime("%d.%m.%Y %H:%M")
        })

        # get components
        res = requests.get(url+'component/project/'+project+'?searchText=&pageSize=99999&pageNumber=1',
            headers=headers, verify=False, timeout=10000)
        text = json.loads(res.text)
        project_info.update({'componentsCount': len(text)})
        for component in text:
            try:
                rec_version = component.get('repositoryMeta').get('latestVersion')
            except Exception: # pylint: disable=broad-exception-caught
                rec_version = ""
            components.update({
                component.get('uuid'): rec_version
            })

        # get vulnerabilities
        res = requests.get(url+'vulnerability/project/'+project,
            headers=headers, verify=False, timeout=10000)
        text = json.loads(res.text)
        project_info.update({'vulnsCount': len(text)})
        for vuln in text:
            component = vuln.get('components')[0]
            group = component.get('group')
            if group is None:
                group = ""
            vuln_id = vuln.get('vulnId')
            vuln_link = RichText()
            if vuln_id.lower().find('cve') != -1:
                vuln_link.add(vuln_id, url_id=doc.build_url_id('https://nvd.nist.gov/vuln/detail/'+vuln_id))
            elif vuln_id.lower().find('ghsa') != -1:
                vuln_link.add(vuln_id, url_id=doc.build_url_id('https://github.com/advisories/'+vuln_id))
            else:
                vuln_link = vuln_id
            vulns.append({
                "component": component.get('name'),
                "version": component.get('version'),
                "rec_version": components.get(component.get('uuid')),
                "group": group,
                "vulnerability": vuln_link,
                "severity": vuln.get('severity') 
            })

        # sort, filter vulnerabilities and components
        uniq_vulns = sorted(list(map(dict, set(tuple(sorted(sub.items())) for sub in vulns))), key=lambda d: (d['component'], d['severity'], d['version']))
        vulns_severity = [i for i in uniq_vulns if (i['severity'].lower() in severities)]
        vuln_component_temp1 = set()
        vuln_component_temp2 = []
        for dic in vulns_severity:
            vuln_component_temp1.add(str(dic.get('component'))+', '+str(dic.get('version'))+', '+str(dic.get('group'))+', '+str(dic.get('rec_version')))
        for v in vuln_component_temp1:
            i = v.split(', ')
            vuln_component_temp2.append({
                "component": i[0],
                "version": i[1],
                "group": i[2],
                "rec_version": i[3]
            })
        vuln_components = sorted(list(map(dict, set(tuple(sorted(sub.items())) for sub in vuln_component_temp2))), key=lambda d: (d['component'], d['version']))

        # render and save result in docx report
        if report_type == "word":
            doc.render({
                'vuln': uniq_vulns,
                'vuln_component': vuln_components,
                'severities': ', '.join(severities),
                'project': project_info
            })
            doc.save("reports/result.docx")
        else:

            # write and save result in excel report
            ws1 = excel["Общая информация"]
            ws1['D2'].value = project_name_str + ' (версия: ' + project_info.get('version') + ')'
            ws1['D2'].hyperlink = url.split('api/v1/')[0]+'projects/'+project
            ws1['D3'] = project_info.get('componentsCount')
            ws1['D4'] = project_info.get('vulnsCount')
            ws1['D5'] = project_info.get('lastBomImport')
            ws1['D6'] = project_info.get('date')
            ws2 = excel["Sheet2"]
            ws2.title =  ', '.join(severities) + " компоненты"
            for num, component in enumerate(vuln_components):
                ws2.cell(row=num+2, column=1, value=num+1)
                ws2.cell(row=num+2, column=2, value=component.get('component'))
                ws2.cell(row=num+2, column=3, value=str(component.get('version')))
                ws2.cell(row=num+2, column=4, value=component.get('group'))
                ws2.cell(row=num+2, column=5, value=str(component.get('rec_version')))
            ws3 = excel["Все срабатывания"]
            for num, vuln in enumerate(uniq_vulns):
                ws3.cell(row=num+2, column=1, value=num+1)
                ws3.cell(row=num+2, column=2, value=vuln.get('component'))
                ws3.cell(row=num+2, column=3, value=str(vuln.get('version')))
                ws3.cell(row=num+2, column=4, value=vuln.get('group'))
                vuln_name = vuln.get('vulnerability')
                if isinstance(vuln_name, RichText):
                    vuln_id = str(vuln_name).split('preserve">')[1].split('</w:t')[0]
                    ws3.cell(row=num+2, column=5, value=vuln_id)
                    if vuln_id.lower().find('cve') != -1:
                        ws3.cell(row=num+2, column=5).hyperlink = 'https://nvd.nist.gov/vuln/detail/'+vuln_id
                    elif vuln_id.lower().find('ghsa') != -1:
                        ws3.cell(row=num+2, column=5).hyperlink = 'https://github.com/advisories/'+vuln_id
                else:
                    ws3.cell(row=num+2, column=5, value=str(vuln_name))
                ws3.cell(row=num+2, column=6, value=vuln.get('severity').lower())
            excel.save("reports/result.xlsx")

        return report_type
    except Exception as e: # pylint: disable=broad-exception-caught
        return e

def get_projects(url, token):
    # header for auth request
    headers = {
        'X-Api-Key': token
    }

    # format url to "protocol"://"domain"/api/v1/
    if not validators.url(url):
        raise ValueError('URL not valid')
    url = url.split('/api/v')[0] + '/api/v1/'
    res = requests.get(url+
        "project?excludeInactive=true&onlyRoot=false&searchText=&sortName=lastBomImport&sortOrder=desc&pageSize=99999&pageNumber=1",
        headers=headers, verify=False, timeout=1000)
    return res.text