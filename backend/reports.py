""" Module for tasks with reports """

import json
import logging
import os
import re
from datetime import datetime

import requests
from docxtpl import DocxTemplate, RichText
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from backend.dependency_graph import compute_graph_levels
from backend.param_validators import (
    check_format_url,
    check_project,
    check_token,
    http_timeout,
    verify_tls,
)

logger = logging.getLogger(__name__)



SUMMARY_SCHEMA_VERSION = 1


def _build_summary(project_name_str, project_url, project_info, vuln_components):
    """ Machine-readable companion to the docx/xlsx report.

    Produced as summary.json inside the ZIP so CI pipelines can post-process
    the report (severity gates, dashboards, ...) without parsing Office files.
    Kept intentionally flat - one component per entry, vulnerabilities nested
    inline - and only exposes JSON-serializable fields.
    """
    return {
        "schemaVersion": SUMMARY_SCHEMA_VERSION,
        "project": {
            "name": project_name_str,
            "version": project_info.get("version"),
            "url": project_url,
            "lastBomImport": project_info.get("lastBomImport"),
            "generatedAt": project_info.get("date"),
            "componentsCount": project_info.get("componentsCount"),
            "vulnerableComponentsCount": project_info.get("vulnComponentsCount"),
            "vulnerabilitiesCount": project_info.get("vulnsCount"),
            "suppressedCount": project_info.get("suppressedCount", 0),
        },
        "components": [
            {
                "name": c.get("name"),
                "version": c.get("version"),
                "group": c.get("group"),
                "lastVersion": c.get("last_version"),
                "isDirectDependency": c.get("is_direct_dependency"),
                "graphLevel": c.get("graph_level"),
                "severity": c.get("severity"),
                "vulnerabilities": [
                    {
                        "id": v.get("id"),
                        "link": v.get("link"),
                        "severity": v.get("severity"),
                        "priority": v.get("priority"),
                        "addInfo": v.get("add_info"),
                        "analysisState": v.get("analysis_state"),
                        "analysisJustification": v.get("analysis_justification"),
                        "analysisResponse": v.get("analysis_response"),
                        "analysisDetail": v.get("analysis_detail"),
                        "isSuppressed": v.get("is_suppressed"),
                        "cvss": v.get("cvss"),
                        "epss": v.get("epss"),
                        "isKev": bool(v.get("is_kev", False)),
                        "isPoc": bool(v.get("is_poc", False)),
                        "isNucleiTemplate": bool(v.get("is_nuclei_template", False)),
                    }
                    for v in c.get("vulnerabilities") or []
                ],
            }
            for c in vuln_components
        ],
    }


def get_severity(severities):
    """ Get level and name of severity by list severities """
    severity = {
        "unknown": 0,
        "undefined": 0,
        "info": 0,
        "low": 1,
        "medium": 2,
        "high": 3,
        "critical": 4
    }
    level = max(severity.get((x or "").lower(), 0) for x in severities)
    return level, [key for key, val in severity.items() if val == level][0]


def _resolve_params(config):
    """ Pull url/token/project out of the request config, validate them """
    raw_url = os.getenv("DT_URL") or (config.get("url") or [""])[0]
    url = check_format_url(raw_url)
    token = os.getenv("DT_TOKEN") or (config.get("token") or [""])[0]
    headers = check_token(token, url)
    project_raw = (config.get("project") or [""])[0]
    # form flow sends "name version (uuid)"; API flow sends a bare UUID
    project_match = re.search(r"\(([^()]+)\)\s*$", project_raw)
    project_id = project_match.group(1) if project_match else project_raw.strip()
    project = check_project(project_id)
    return url, headers, project


def _fetch_project_info(url, headers, project):
    """ Read project metadata from DT and shape it for the report """
    logger.info("Fetching project metadata")
    res = requests.get(url+"project/"+project, headers=headers,
                       verify=verify_tls(), timeout=http_timeout())
    res.raise_for_status()
    text = res.json()
    project_name_str = text.get("name")
    metrics = text.get("metrics") or {}
    last_bom_ms = int(text.get("lastBomImport") or 0)
    project_info = {
        # Plain string here; the docx render step decorates it as a hyperlink
        "name": project_name_str,
        "version": text.get("version") or "no version",
        "lastBomImport": datetime.fromtimestamp(last_bom_ms / 1000).strftime(
            "%d.%m.%Y %H:%M"),
        # Raw epoch (ms) for chronological comparisons (e.g. diff A/B swap).
        # Stays out of the user-facing report.
        "lastBomImportTimestamp": last_bom_ms,
        "date": datetime.now().strftime("%d.%m.%Y %H:%M"),
        "componentsCount": metrics.get("components"),
        "vulnsCount": metrics.get("vulnerabilities"),
        "vulnComponentsCount": metrics.get("vulnerableComponents"),
    }
    logger.debug(f"Project info retrieved: {project_info}")
    if text.get("directDependencies"):
        direct_uuids = [x.get("uuid")
                        for x in json.loads(text.get("directDependencies"))]
    else:
        direct_uuids = []
    return project_info, project_name_str, direct_uuids


def _fetch_sbom(url, headers, project):
    """ Pull the CycloneDX SBOM (vulnerabilities + dependency graph) """
    logger.info("Fetching SBOM with vulnerabilities")
    res = requests.get(url+"bom/cyclonedx/project/"+project
                       +"?format=json&variant=withVulnerabilities&download=true",
                       headers=headers, verify=verify_tls(), timeout=http_timeout())
    res.raise_for_status()
    text = res.json()
    vulnerabilities = text.get("vulnerabilities") or []
    deps_deps = {}
    for deps in text.get("dependencies") or []:
        deps_deps[deps.get("ref")] = deps.get("dependsOn")
    return vulnerabilities, deps_deps


def _fetch_findings(url, headers, project):
    """ Pull VEX analysis from the findings API.

    The CycloneDX variant=withVulnerabilities does not always carry the
    analysis block (depends on DT version and how the VEX was imported);
    findings is the source of truth for the audit state shown in the DT UI.
    Returns a (vuln_id, component_uuid) -> analysis dict map.
    """
    logger.info("Fetching VEX analysis from findings API")
    res = requests.get(url+"finding/project/"+project+"?suppressed=true",
                       headers=headers, verify=verify_tls(), timeout=http_timeout())
    res.raise_for_status()
    analysis_by_pair = {}
    for finding in res.json() or []:
        vuln_id = (finding.get("vulnerability") or {}).get("vulnId")
        component_uuid = (finding.get("component") or {}).get("uuid")
        analysis = finding.get("analysis") or {}
        if vuln_id and component_uuid:
            analysis_by_pair[(vuln_id, component_uuid)] = {
                "state": (analysis.get("state") or "").lower(),
                "justification": analysis.get("justification") or "",
                "is_suppressed": bool(analysis.get("isSuppressed")),
            }
    logger.info(f"Findings API returned analysis for "
                f"{len(analysis_by_pair)} (vuln, component) pairs")
    return analysis_by_pair


def _fetch_components(url, headers, project, direct_uuids, deps_deps):
    """ Pull every component of the project and shape it for the report """
    res = requests.get(url+"component/project/"+project+
        "?searchText=&pageSize=99999&pageNumber=1",
        headers=headers, verify=verify_tls(), timeout=http_timeout())
    res.raise_for_status()
    components = {}
    for component in res.json():
        try:
            last_version = component.get("repositoryMeta").get("latestVersion")
        except AttributeError:
            last_version = ""
        components[component.get("uuid")] = {
            "name": component.get("name"),
            "version": component.get("version"),
            "group": component.get("group") or "",
            "last_version": last_version,
            "is_direct_dependency": component.get("uuid") in direct_uuids,
            "dependencies": deps_deps[component.get("uuid")],
            "vulnerabilities": [],
            "severity": "",
            "severity_level": 0,
            "graph_level": None,
        }
    logger.info(f"{len(components)} components processed")
    return components


_SUPPRESSED_STATES = {"resolved", "resolved_with_pedigree",
                      "false_positive", "not_affected"}


_CVE_ID_PATTERN = re.compile(r"CVE-\d{4}-\d{4,7}", re.IGNORECASE)
_CVEPAAS_BATCH_SIZE = 50


def _canonical_cve_ids(vulnerabilities):
    """ Unique canonical CVE ids that are safe to feed to CVE-PaaS """
    ids = set()
    for vuln in vulnerabilities:
        vuln_id = vuln.get("id") or ""
        if _CVE_ID_PATTERN.fullmatch(vuln_id):
            ids.add(vuln_id)
    return ids


def _fetch_cve_paas(cve_ids):
    """ Batch-fetch CVE-PaaS enrichment for the given CVE ids.

    Returns {cve_id: data_dict}. CVE-PaaS exposes POST /v1/cve which
    accepts up to 50 ids per call, so we chunk. Any failure (network,
    non-2xx, malformed JSON) is logged and skipped - the report is
    rendered without enrichment for that batch instead of being
    aborted (graceful degradation).
    """
    base_url = (os.getenv("CVEPAAS_URL") or "").rstrip("/")
    if not base_url or not cve_ids:
        return {}
    headers = {}
    key = os.getenv("DTRG_CVEPAAS_KEY")
    if key:
        headers["X-API-Key"] = key
    result = {}
    cve_list = sorted(cve_ids)
    for i in range(0, len(cve_list), _CVEPAAS_BATCH_SIZE):
        batch = cve_list[i:i + _CVEPAAS_BATCH_SIZE]
        try:
            res = requests.post(
                base_url + "/v1/cve",
                json={"cve_ids": batch},
                headers=headers,
                verify=verify_tls(),
                timeout=http_timeout(),
            )
            res.raise_for_status()
            payload = res.json()
            if isinstance(payload, dict):
                result.update(payload)
        except (requests.RequestException, ValueError) as e:
            logger.warning(
                f"CVE-PaaS batch fetch failed ({len(batch)} ids): {e}; "
                f"continuing without enrichment for this batch"
            )
            continue
    logger.info(f"CVE-PaaS returned enrichment for {len(result)} of "
                f"{len(cve_list)} requested ids")
    return result


def _attach_vulnerabilities(components, vulnerabilities, analysis_by_pair,
                            cve_paas_data, doc=None):
    """ Distribute vulnerabilities across components, with VEX + CVE-PaaS data.

    For every (vuln, affected component) pair we resolve the analysis state
    (findings API wins, SBOM block is the fallback for response/detail),
    build the per-vuln link, look up CVE-PaaS enrichment from the
    pre-fetched cve_paas_data dict, and append the result to the component's
    vulnerabilities list. The doc handle is needed for docxtpl's hyperlink
    builder; pass None when only xlsx/json output is wanted (e.g. diff
    reports) and word_link will fall back to the plain id.
    """
    logger.info("Processing component vulnerabilities")
    for vuln in vulnerabilities:
        sbom_analysis = vuln.get("analysis") or {}
        sbom_state = (sbom_analysis.get("state") or "").lower()
        sbom_justification = sbom_analysis.get("justification") or ""
        analysis_response = ", ".join(sbom_analysis.get("response") or [])
        analysis_detail = sbom_analysis.get("detail") or ""
        for component in vuln.get("affects"):
            vuln_id = vuln.get("id")
            component_ref = component.get("ref")
            # Findings API wins over the SBOM analysis block; SBOM is
            # the fallback for response/detail (findings does not expose those).
            finding_analysis = analysis_by_pair.get((vuln_id, component_ref), {})
            analysis_state = finding_analysis.get("state") or sbom_state
            analysis_justification = (finding_analysis.get("justification")
                                      or sbom_justification)
            is_suppressed = (
                finding_analysis.get("is_suppressed", False)
                or analysis_state in _SUPPRESSED_STATES
            )
            if "cve" in vuln_id.lower():
                vuln_link = "https://nvd.nist.gov/vuln/detail/"+vuln_id
                # only canonical CVE-YYYY-NNNN ids have an entry in cve_paas_data
                cve_id = vuln_id if _CVE_ID_PATTERN.fullmatch(vuln_id) else ""
            elif "ghsa" in vuln_id.lower():
                vuln_link = "https://github.com/advisories/"+vuln_id
# https://docs.github.com/en/rest/security-advisories/global-advisories?apiVersion=2022-11-28
                cve_id = ""
            else:
                vuln_link = vuln_id
                cve_id = ""
            if doc is not None and vuln_link != vuln_id:
                vuln_word_link = RichText()
                vuln_word_link.add(vuln_id, url_id=doc.build_url_id(vuln_link))
            else:
                vuln_word_link = vuln_id
            severity_level, severity = get_severity(list(x.get("severity")
                                                         for x in vuln.get("ratings")))
            cve_paas = cve_paas_data.get(cve_id, {}) if cve_id else {}
            details = cve_paas.get("Details") or {}
            links = details.get("Links") or {}
            is_kev = bool(details.get("is_exploited"))
            is_poc = bool(details.get("is_poc"))
            is_template = bool(details.get("is_template"))
            # Surface KEV / POC / Nuclei links in add_info regardless of
            # priority - the boolean flags from CVE-PaaS already gate the
            # presence of useful URLs, no need to also gate on Priority.
            add_info = []
            if is_kev:
                kev_link = links.get("CISA KEV") or {}
                kev_url = kev_link.get("url")
                if kev_url:
                    add_info.append(f"KEV: {kev_url}")
            if is_poc:
                for poc in links.get("POC") or []:
                    poc_url = (poc or {}).get("url")
                    if poc_url:
                        add_info.append(f"POC: {poc_url}")
            if is_template:
                nuclei = links.get("Nuclei templates") or {}
                template_url = nuclei.get("template_url")
                if template_url:
                    add_info.append(f"Nuclei: {template_url}")
            components[component_ref]["vulnerabilities"].append({
                "uuid": vuln.get("bom-ref"),
                "id": vuln_id,
                "link": vuln_link,
                "word_link": vuln_word_link,
                "severity": severity,
                "severity_level": severity_level,
                "priority": cve_paas.get("Priority") or severity,
                "add_info": ", ".join(sorted(set(add_info))),
                "analysis_state": analysis_state,
                "analysis_justification": analysis_justification,
                "analysis_response": analysis_response,
                "analysis_detail": analysis_detail,
                "is_suppressed": is_suppressed,
                "cvss": details.get("CVSS"),
                "epss": details.get("EPSS"),
                "is_kev": is_kev,
                "is_poc": is_poc,
                "is_nuclei_template": is_template,
            })
    logger.info("Vulnerabilities assigned to components")


def _filter_suppressed(components, project_info):
    """ Drop VEX-suppressed vulns unless DTRG_INCLUDE_SUPPRESSED is set.

    Mutates components in place. Records the count of dropped (or kept,
    when included) findings on project_info["suppressedCount"].
    """
    include_suppressed = os.getenv("DTRG_INCLUDE_SUPPRESSED",
                                   "false").lower() in ["true", "1", "t"]
    suppressed_count = 0
    for value in components.values():
        kept = []
        for vuln in value["vulnerabilities"]:
            if vuln["is_suppressed"]:
                suppressed_count += 1
                if not include_suppressed:
                    continue
            kept.append(vuln)
        value["vulnerabilities"] = kept
    project_info["suppressedCount"] = suppressed_count
    logger.info(f"VEX-suppressed vulnerabilities: {suppressed_count} "
                f"(included in report: {include_suppressed})")


def _compute_severity(components):
    """ Roll up the per-component severity from its remaining vulns """
    logger.info("Computing final severity levels for components")
    use_priority = bool(os.getenv("CVEPAAS_URL"))
    for value in components.values():
        vulns = value.get("vulnerabilities")
        if not vulns:
            continue
        if use_priority:
            severity_level, severity = get_severity(list(x.get("priority").lower()
                                                         for x in vulns))
        else:
            severity_level, severity = get_severity(list(x.get("severity")
                                                         for x in vulns))
        value["severity"] = severity
        value["severity_level"] = severity_level


def _sort_vulnerable(components):
    """ Return components that have vulns, sorted high-severity first """
    vuln_components = {k: v for k, v in components.items() if v.get("vulnerabilities")}
    return list(dict(sorted(vuln_components.items(),
                            key=lambda item: item[1]["severity_level"],
                            reverse=True)).values())


def _render_docx(doc, project_info, project_name_str, project_url,
                 vuln_components, output_dir):
    """ Render the Word report into output_dir/result.docx """
    logger.info("Generating Word report")
    # decorate the project name with a hyperlink only at render time, so the
    # data layer above stays free of docxtpl-specific objects
    project_for_render = dict(project_info)
    project_link = RichText()
    project_link.add(project_name_str, url_id=doc.build_url_id(project_url))
    project_for_render["name"] = project_link
    doc.render({
        "project": project_for_render,
        "components": vuln_components,
    })
    doc.save(os.path.join(output_dir, "result.docx"))
    logger.info("Word report saved")


def _render_xlsx(excel, project_info, project_name_str, project_url,
                 vuln_components, output_dir):
    """ Render the Excel report into output_dir/result.xlsx """
    logger.info("Generating Excel report")
    ws1 = excel["General information"]
    ws1["D2"].value = (project_name_str + " (version: "
                       + project_info.get("version") + ")")
    ws1["D2"].hyperlink = project_url
    ws1["D3"] = project_info.get("componentsCount")
    ws1["D4"] = project_info.get("vulnsCount")
    ws1["D5"] = project_info.get("vulnComponentsCount")
    ws1["D6"] = project_info.get("lastBomImport")
    ws1["D7"] = project_info.get("date")
    if not vuln_components:
        del excel["Vulnerable dependencies"]
        del excel["All issues"]
        excel.save(os.path.join(output_dir, "result.xlsx"))
        logger.info("Excel report saved")
        return
    ws2 = excel["Vulnerable dependencies"]
    ws3 = excel["All issues"]
    vuln_num = 0
    for num, component in enumerate(vuln_components):
        ws2.cell(row=num+2, column=1, value=num+1)
        ws2.cell(row=num+2, column=2, value=component.get("name"))
        ws2.cell(row=num+2, column=3, value=str(component.get("version")))
        ws2.cell(row=num+2, column=4, value=component.get("group"))
        if component.get("is_direct_dependency"):
            final_severity = f"{str(component.get('severity'))} in direct dependency"
        else:
            final_severity = str(component.get("severity"))
        ws2.cell(row=num+2, column=5, value=final_severity)
        ws2.cell(row=num+2, column=6, value=str(component.get("last_version")))
        graph_level = component.get("graph_level")
        ws2.cell(row=num+2, column=7,
                 value="" if graph_level is None else graph_level)
        for vuln in component.get("vulnerabilities"):
            ws3.cell(row=num+2+vuln_num, column=1, value=num+1+vuln_num)
            ws3.cell(row=num+2+vuln_num, column=2, value=vuln.get("id"))
            if isinstance(vuln.get("word_link"), RichText):
                ws3.cell(row=num+2+vuln_num, column=2).hyperlink = vuln.get("link")
            ws3.cell(row=num+2+vuln_num, column=3, value=vuln.get("severity"))
            ws3.cell(row=num+2+vuln_num, column=4,
                     value=(vuln.get("priority") or "").lower())
            ws3.cell(row=num+2+vuln_num, column=5, value=component.get("name"))
            ws3.cell(row=num+2+vuln_num, column=6, value=component.get("version"))
            ws3.cell(row=num+2+vuln_num, column=7, value=vuln.get("add_info"))
            ws3.cell(row=num+2+vuln_num, column=7).alignment = Alignment(wrap_text=True)
            ws3.cell(row=num+2+vuln_num, column=8,
                     value=vuln.get("analysis_state") or "")
            vuln_num += 1
        vuln_num -= 1
    excel.save(os.path.join(output_dir, "result.xlsx"))
    logger.info("Excel report saved")


def _render_summary(project_name_str, project_url, project_info,
                    vuln_components, output_dir):
    """ Write the JSON summary alongside the docx/xlsx """
    summary = _build_summary(project_name_str, project_url,
                             project_info, vuln_components)
    with open(os.path.join(output_dir, "summary.json"), "w",
              encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    logger.info("JSON summary saved")


def compute_diff(data_a, data_b):
    """ Compute added / removed / common vulnerabilities between two snapshots.

    Each entry in vuln_components carries a name, group and a list of vulns.
    Identity for matching is (component_name, component_group, vuln_id) -
    component version is part of the value, not the key, so a CVE whose host
    component got upgraded between A and B (but whose id is unchanged) shows
    up under "common" with both versions visible. data_a and data_b are
    _load_project return values.

    Returns {"added": [...], "removed": [...], "common": [...]}.
    """
    def _index(data):
        idx = {}
        for component in data["vuln_components"]:
            for vuln in component["vulnerabilities"]:
                key = (component.get("name"), component.get("group") or "",
                       vuln.get("id"))
                idx[key] = (component, vuln)
        return idx

    def _entry(component, vuln):
        return {
            "component": component.get("name"),
            "group": component.get("group") or "",
            "componentVersion": component.get("version"),
            "vulnerability": vuln.get("id"),
            "link": vuln.get("link"),
            "severity": vuln.get("severity"),
            "priority": vuln.get("priority"),
            "addInfo": vuln.get("add_info") or "",
            "analysisState": vuln.get("analysis_state") or "",
            "isSuppressed": bool(vuln.get("is_suppressed", False)),
            "cvss": vuln.get("cvss"),
            "epss": vuln.get("epss"),
            "isKev": bool(vuln.get("is_kev", False)),
            "isPoc": bool(vuln.get("is_poc", False)),
            "isNucleiTemplate": bool(vuln.get("is_nuclei_template", False)),
        }

    idx_a = _index(data_a)
    idx_b = _index(data_b)
    keys_a = set(idx_a.keys())
    keys_b = set(idx_b.keys())

    added = [_entry(*idx_b[k]) for k in sorted(keys_b - keys_a)]
    removed = [_entry(*idx_a[k]) for k in sorted(keys_a - keys_b)]
    common = []
    for key in sorted(keys_a & keys_b):
        ca, va = idx_a[key]
        cb, vb = idx_b[key]
        common.append({
            "component": ca.get("name"),
            "group": ca.get("group") or "",
            "componentVersionA": ca.get("version"),
            "componentVersionB": cb.get("version"),
            "vulnerability": va.get("id"),
            "link": vb.get("link") or va.get("link"),
            "severity": vb.get("severity"),
            "priority": vb.get("priority"),
            "addInfo": vb.get("add_info") or va.get("add_info") or "",
            "analysisStateA": va.get("analysis_state") or "",
            "analysisStateB": vb.get("analysis_state") or "",
            "isSuppressedA": bool(va.get("is_suppressed", False)),
            "isSuppressedB": bool(vb.get("is_suppressed", False)),
            "cvss": vb.get("cvss") or va.get("cvss"),
            "epss": vb.get("epss") or va.get("epss"),
            "isKev": bool(vb.get("is_kev", False) or va.get("is_kev", False)),
            "isPoc": bool(vb.get("is_poc", False) or va.get("is_poc", False)),
            "isNucleiTemplate": bool(vb.get("is_nuclei_template", False)
                                     or va.get("is_nuclei_template", False)),
        })
    return {"added": added, "removed": removed, "common": common}


def _load_project(url, headers, project, doc=None):
    """ Run the full per-project pipeline and return everything renderers need.

    Stitches together the four DT fetches and the four processing helpers
    so callers (single-project create_report, two-project diff) do not
    repeat the orchestration. doc is forwarded to _attach_vulnerabilities
    for RichText hyperlinks; pass None when no docx output is needed.

    Returns a dict with keys: info, name, url, components, vuln_components.
    """
    project_info, project_name_str, direct_uuids = _fetch_project_info(
        url, headers, project)
    vulnerabilities, deps_deps = _fetch_sbom(url, headers, project)
    analysis_by_pair = _fetch_findings(url, headers, project)
    components = _fetch_components(url, headers, project,
                                   direct_uuids, deps_deps)
    cve_paas_data = _fetch_cve_paas(_canonical_cve_ids(vulnerabilities))
    _attach_vulnerabilities(components, vulnerabilities, analysis_by_pair,
                            cve_paas_data, doc)
    _filter_suppressed(components, project_info)
    _compute_severity(components)
    vuln_components = _sort_vulnerable(components)
    logger.info(f"{len(vuln_components)} vulnerable components found")
    compute_graph_levels(components)
    return {
        "info": project_info,
        "name": project_name_str,
        "url": url.split("api/v1/")[0] + "projects/" + project,
        "components": components,
        "vuln_components": vuln_components,
    }


def create_report(config, output_dir):
    """ Create report from DT into the per-request output_dir """
    logger.info("Report generation started")
    doc = DocxTemplate("reports/draft.docx")
    excel = load_workbook("reports/draft.xlsx")

    try:
        url, headers, project = _resolve_params(config)
        data = _load_project(url, headers, project, doc=doc)

        _render_docx(doc, data["info"], data["name"], data["url"],
                     data["vuln_components"], output_dir)
        _render_xlsx(excel, data["info"], data["name"], data["url"],
                     data["vuln_components"], output_dir)
        _render_summary(data["name"], data["url"], data["info"],
                        data["vuln_components"], output_dir)

        report_name = data["name"] or project
        return (f"{report_name} {data['info'].get('version')} "
                f"({datetime.now().strftime('%d.%m.%Y')})", data["components"])
    except (ValueError, ConnectionError) as e:
        logger.error(f"Error while generating report: {e}")
        return e, []


def _project_summary(data):
    """ Trim a _load_project dict down to the JSON-serializable bits """
    info = data["info"]
    return {
        "name": data["name"],
        "version": info.get("version"),
        "url": data["url"],
        "lastBomImport": info.get("lastBomImport"),
        "componentsCount": info.get("componentsCount"),
        "vulnerableComponentsCount": info.get("vulnComponentsCount"),
        "vulnerabilitiesCount": info.get("vulnsCount"),
        "suppressedCount": info.get("suppressedCount", 0),
    }


def _fill_diff_general(ws, data_a, data_b):
    """ Fill the side-by-side metadata sheet of draft_diff.xlsx.

    Template provides the row labels in column B and the A / B header
    cells in D2 / E2; values land in columns D (A) and E (B).
    """
    def _write(col, data):
        info = data["info"]
        ws.cell(row=3, column=col,
                value=f"{data['name']} (version: {info.get('version')})").hyperlink = \
            data["url"]
        ws.cell(row=4, column=col, value=info.get("componentsCount"))
        ws.cell(row=5, column=col, value=info.get("vulnsCount"))
        ws.cell(row=6, column=col, value=info.get("vulnComponentsCount"))
        ws.cell(row=7, column=col, value=info.get("lastBomImport"))
        ws.cell(row=8, column=col, value=info.get("date"))
    _write(4, data_a)
    _write(5, data_b)


def _fill_diff_vulnerable_dependencies(ws, data_a, data_b):
    """ Union of vulnerable components from A and B, one row per (name, group).

    Per the user's design decisions:
    - Final severity is B's (current state of the newer project)
    - Last version / Graph level: B's, falling back to A when the component
      is gone in B (it was removed between snapshots)
    """
    def _index(vuln_components):
        return {(c.get("name"), c.get("group") or ""): c for c in vuln_components}

    idx_a = _index(data_a["vuln_components"])
    idx_b = _index(data_b["vuln_components"])
    all_keys = sorted(set(idx_a) | set(idx_b))
    for num, key in enumerate(all_keys):
        ca = idx_a.get(key)
        cb = idx_b.get(key)
        primary = cb or ca
        row = num + 2
        ws.cell(row=row, column=1, value=num + 1)
        ws.cell(row=row, column=2, value=primary.get("name"))
        ws.cell(row=row, column=3, value=primary.get("group"))
        ws.cell(row=row, column=4, value=str(ca.get("version")) if ca else "")
        ws.cell(row=row, column=5, value=str(cb.get("version")) if cb else "")
        severity = primary.get("severity") or ""
        if primary.get("is_direct_dependency"):
            severity = f"{severity} in direct dependency"
        ws.cell(row=row, column=6, value=severity)
        ws.cell(row=row, column=7, value=str(primary.get("last_version") or ""))
        graph_level = primary.get("graph_level")
        ws.cell(row=row, column=8,
                value="" if graph_level is None else graph_level)


def _fill_diff_issues_sheet(ws, entries, version_kind):
    """ Fill one of the three issue sheets in draft_diff.xlsx.

    version_kind = "single" for Added / Removed (each entry has one
    componentVersion) or "compare" for Common ("oldVersion → newVersion"
    when they differ, otherwise just the version).
    """
    for num, entry in enumerate(entries):
        row = num + 2
        ws.cell(row=row, column=1, value=num + 1)
        cell = ws.cell(row=row, column=2, value=entry["vulnerability"])
        link = entry.get("link") or ""
        if link.startswith("http"):
            cell.hyperlink = link
        ws.cell(row=row, column=3, value=entry["severity"])
        ws.cell(row=row, column=4, value=(entry.get("priority") or "").lower())
        ws.cell(row=row, column=5, value=entry["component"])
        if version_kind == "compare":
            v_a = entry.get("componentVersionA")
            v_b = entry.get("componentVersionB")
            version = f"{v_a} → {v_b}" if v_a != v_b else str(v_b or "")
        else:
            version = str(entry.get("componentVersion") or "")
        ws.cell(row=row, column=6, value=version)
        ws.cell(row=row, column=7, value=entry.get("addInfo") or "")
        ws.cell(row=row, column=7).alignment = Alignment(wrap_text=True)
        # Common rows use B's analysis state (current); Added/Removed have
        # a single analysisState already.
        state = entry.get("analysisStateB") or entry.get("analysisState") or ""
        ws.cell(row=row, column=8, value=state)


def _render_diff_xlsx(diff, data_a, data_b, output_dir):
    """ Fill reports/draft_diff.xlsx with the diff and save into output_dir """
    logger.info("Generating diff Excel report")
    excel = load_workbook("reports/draft_diff.xlsx")
    _fill_diff_general(excel["General information"], data_a, data_b)
    _fill_diff_vulnerable_dependencies(
        excel["Vulnerable dependencies"], data_a, data_b)
    _fill_diff_issues_sheet(excel["Added issues"], diff["added"], "single")
    _fill_diff_issues_sheet(excel["Removed issues"], diff["removed"], "single")
    _fill_diff_issues_sheet(excel["Common issues"], diff["common"], "compare")
    excel.save(os.path.join(output_dir, "result.xlsx"))
    logger.info("Diff Excel report saved")


def _render_diff_summary(diff, data_a, data_b, output_dir):
    """ Write the JSON summary for a diff report """
    summary = {
        "schemaVersion": SUMMARY_SCHEMA_VERSION,
        "kind": "diff",
        "generatedAt": datetime.now().strftime("%d.%m.%Y %H:%M"),
        "projectA": _project_summary(data_a),
        "projectB": _project_summary(data_b),
        "diff": diff,
    }
    with open(os.path.join(output_dir, "summary.json"), "w",
              encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    logger.info("Diff JSON summary saved")


def create_diff_report(config_a, config_b, output_dir):
    """ Generate a diff report between two DT projects.

    config_a / config_b have the same shape as create_report's config
    (url / token / project), so the route layer can just split a single
    request into two configs (same URL, same token, two project IDs).
    Returns (report_name, None) on success, (exception, None) on failure -
    None matches create_report's "no graph for this report" signal.
    """
    logger.info("Diff report generation started")
    try:
        url_a, headers_a, project_a = _resolve_params(config_a)
        url_b, headers_b, project_b = _resolve_params(config_b)
        data_a = _load_project(url_a, headers_a, project_a)
        data_b = _load_project(url_b, headers_b, project_b)
        # Ensure A is the older snapshot and B is the newer one, regardless of
        # the order the caller passed them in. We compare DT's lastBomImport
        # timestamp (epoch ms). Equal or missing timestamps leave the order
        # alone.
        ts_a = data_a["info"].get("lastBomImportTimestamp") or 0
        ts_b = data_b["info"].get("lastBomImportTimestamp") or 0
        if ts_a > ts_b:
            logger.info("Swapping A/B so the newer project is B "
                        f"(ts_a={ts_a} > ts_b={ts_b})")
            data_a, data_b = data_b, data_a
            project_a, project_b = project_b, project_a
        diff = compute_diff(data_a, data_b)
        logger.info(f"Diff computed: +{len(diff['added'])} added, "
                    f"-{len(diff['removed'])} removed, "
                    f"={len(diff['common'])} common")
        _render_diff_xlsx(diff, data_a, data_b, output_dir)
        _render_diff_summary(diff, data_a, data_b, output_dir)
        name_a = data_a["name"] or project_a
        name_b = data_b["name"] or project_b
        ver_a = data_a["info"].get("version") or ""
        ver_b = data_b["info"].get("version") or ""
        return (f"diff {name_a} {ver_a} vs {name_b} {ver_b} "
                f"({datetime.now().strftime('%d.%m.%Y')})", None)
    except (ValueError, ConnectionError) as e:
        logger.error(f"Error while generating diff report: {e}")
        return e, None
